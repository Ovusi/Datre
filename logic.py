import pandas as pd
from flask_admin import Admin
from flask_login import LoginManager, UserMixin
from flask_security import Security
from flask import request
from database import *
from openpyxl import load_workbook


def xlsx(file):
    workbook = load_workbook(filename=file)
    workbook.sheetnames()
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
        pass
    for col in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
        pass


def csv(file):
    pass


def upload():
    file = request.files['']
    new_file = Charts(name=file.filename, data=file.read())
    db.session.add(new_file)
    db.session.commit()


def login():
    pass


def logout():
    pass


def admin():
    pass


def download():
    pass
