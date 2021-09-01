from database.database import *
from openpyxl import load_workbook


""" This module contains functions for file handling and manipulation.
    """


def xlsx(file):
    workbook = load_workbook(filename=file)
    workbook.sheetnames()
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
        pass
    for col in sheet.iter_cols(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
        pass


def csv(file):
    pass


def upload(file, database):
    new_file = Charts(name=file.filename, data=file.read())
    database.session.add(new_file)
    database.session.commit()


def download():
    pass
